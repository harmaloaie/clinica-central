#!/usr/bin/env python3
"""
Reconstruieste assets/data/analize.json din fisierele Excel
din folderul data-source/.

Folosire:
    python3 scripts/build_data.py

Pentru fiecare fisier .xlsx din data-source/ scriptul cauta
configurarea in scripts/labs_config.yaml. Daca nu o gaseste,
incearca auto-detectie (cauta o coloana de denumire + una de pret).
"""

import json
import sys
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("EROARE: openpyxl nu este instalat. Ruleaza: pip install openpyxl pyyaml", file=sys.stderr)
    sys.exit(1)

try:
    import yaml
except ImportError:
    print("EROARE: pyyaml nu este instalat. Ruleaza: pip install openpyxl pyyaml", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
SOURCE_DIR = ROOT / "data-source"
OUTPUT = ROOT / "assets" / "data" / "analize.json"
CONFIG = ROOT / "scripts" / "labs_config.yaml"


def parse_price(value):
    """Converteste o celula in pret float. Returneaza None daca nu e parsabil."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    # Elimina prefix ^^ folosit de Solomed
    s = s.lstrip("^").strip()
    # Inlocuieste virgula cu punct (format romanesc: 160,00)
    s = s.replace(",", ".")
    # Elimina " RON", spatii, etc.
    s = re.sub(r"[^\d.\-]", "", s)
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def clean_text(value):
    """Curata un text de spatii non-breaking si prefixe de tip ^^."""
    if value is None:
        return None
    s = str(value).strip()
    s = s.lstrip("^").strip()
    # Elimina nbsp
    s = s.replace("\u00a0", " ")
    # Colapseaza spatii multiple
    s = re.sub(r"\s+", " ", s)
    return s if s else None


def get_sheet(wb, spec):
    """Returneaza ws-ul pe baza spec-ului (string nume sau int index)."""
    if isinstance(spec, int):
        names = wb.sheetnames
        if 0 <= spec < len(names):
            return wb[names[spec]]
        raise KeyError(f"Sheet index {spec} nu exista. Sheets disponibile: {names}")
    if spec in wb.sheetnames:
        return wb[spec]
    raise KeyError(f"Sheet '{spec}' nu exista. Sheets disponibile: {wb.sheetnames}")


def extract_records(ws, lab_name, lab_cfg):
    """Extrage inregistrarile dintr-un worksheet conform configurarii."""
    header_row = int(lab_cfg.get("header_row", 1))

    # Stabileste indecsii coloanelor (1-indexed) pentru fiecare rol
    col_idx = {}  # role -> column index

    if "columns_by_position" in lab_cfg:
        for pos, role in lab_cfg["columns_by_position"].items():
            col_idx[role] = int(pos)
    elif "columns" in lab_cfg:
        # Citeste header-ul si mapeaza nume -> index
        header_cells = [c.value for c in ws[header_row]]
        for col_name, role in lab_cfg["columns"].items():
            try:
                idx = header_cells.index(col_name) + 1
                col_idx[role] = idx
            except ValueError:
                # Incearca match dupa strip+lowercase
                target = str(col_name).strip().lower()
                found = False
                for i, h in enumerate(header_cells):
                    if h is not None and str(h).strip().lower() == target:
                        col_idx[role] = i + 1
                        found = True
                        break
                if not found:
                    raise KeyError(
                        f"Coloana '{col_name}' nu a fost gasita in header. "
                        f"Header gasit: {header_cells}"
                    )
    else:
        raise ValueError(f"Config lipseste 'columns' sau 'columns_by_position' pentru {lab_name}")

    if "denumire" not in col_idx or "pret" not in col_idx:
        raise ValueError(f"Configul pentru {lab_name} trebuie sa aiba 'denumire' si 'pret'")

    records = []
    skipped = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        denumire = clean_text(row[col_idx["denumire"] - 1]) if col_idx["denumire"] - 1 < len(row) else None
        pret = parse_price(row[col_idx["pret"] - 1]) if col_idx["pret"] - 1 < len(row) else None

        if not denumire or pret is None or pret <= 0:
            skipped += 1
            continue

        rec = {
            "Laborator": lab_name,
            "Denumire": denumire,
            "Categorie": "N/A",
            "Timp": "N/A",
            "Pret": pret,
        }

        if "categorie" in col_idx and col_idx["categorie"] - 1 < len(row):
            cat = clean_text(row[col_idx["categorie"] - 1])
            if cat:
                rec["Categorie"] = cat

        if "timp" in col_idx and col_idx["timp"] - 1 < len(row):
            t = clean_text(row[col_idx["timp"] - 1])
            if t:
                rec["Timp"] = t

        records.append(rec)

    return records, skipped


def auto_detect_config(wb):
    """Pentru un fisier necunoscut, incearca sa ghicesti structura."""
    ws = wb[wb.sheetnames[0]]
    # Cauta randul de header (primele 10 randuri)
    for header_row in range(1, min(11, ws.max_row + 1)):
        cells = [str(c.value).strip().lower() if c.value else "" for c in ws[header_row]]
        denumire_idx = None
        pret_idx = None
        for i, h in enumerate(cells):
            if denumire_idx is None and any(k in h for k in ["denumire", "analiza", "serviciu", "nume"]):
                denumire_idx = i + 1
            if pret_idx is None and any(k in h for k in ["pret", "preț", "price", "cost", "tarif"]):
                pret_idx = i + 1
        if denumire_idx and pret_idx:
            return {
                "sheet": 0,
                "header_row": header_row,
                "columns_by_position": {denumire_idx: "denumire", pret_idx: "pret"},
            }
    return None


def process_file(xlsx_path, config_entry):
    """Proceseaza un singur fisier Excel."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    if config_entry is None:
        # Auto-detect
        print(f"  [{xlsx_path.name}] Nu exista config — incerc auto-detect...")
        auto = auto_detect_config(wb)
        if auto is None:
            print(f"  [{xlsx_path.name}] AUTO-DETECT ESUAT. Adauga un bloc in labs_config.yaml.")
            return [], 0
        # Numele laboratorului = numele fisierului (fara extensie)
        lab_name = xlsx_path.stem
        cfg = {**auto, "lab_name": lab_name}
        print(f"  [{xlsx_path.name}] Auto-detect OK: header_row={auto['header_row']}, "
              f"denumire=col{[k for k,v in auto['columns_by_position'].items() if v=='denumire'][0]}, "
              f"pret=col{[k for k,v in auto['columns_by_position'].items() if v=='pret'][0]}")
    else:
        cfg = config_entry
        lab_name = cfg["lab_name"]

    # Multi-sheet sau single
    sheets_spec = cfg.get("sheets", [cfg.get("sheet", 0)])
    if not isinstance(sheets_spec, list):
        sheets_spec = [sheets_spec]

    all_records = []
    total_skipped = 0
    for sheet_spec in sheets_spec:
        ws = get_sheet(wb, sheet_spec)
        records, skipped = extract_records(ws, lab_name, cfg)
        all_records.extend(records)
        total_skipped += skipped

    return all_records, total_skipped


def main():
    if not SOURCE_DIR.exists():
        print(f"EROARE: folderul {SOURCE_DIR} nu exista", file=sys.stderr)
        sys.exit(1)

    if CONFIG.exists():
        with open(CONFIG, encoding="utf-8") as f:
            config_data = yaml.safe_load(f) or {}
        config_by_file = {entry["file"]: entry for entry in config_data.get("labs", [])}
    else:
        print(f"AVERTIZARE: {CONFIG} nu exista — toate fisierele vor folosi auto-detect")
        config_by_file = {}

    xlsx_files = sorted(SOURCE_DIR.glob("*.xlsx"))
    if not xlsx_files:
        print(f"EROARE: niciun fisier .xlsx in {SOURCE_DIR}", file=sys.stderr)
        sys.exit(1)

    print(f"\n=== Procesez {len(xlsx_files)} fisiere din {SOURCE_DIR.name}/ ===\n")

    all_records = []
    summary = []

    for xlsx_path in xlsx_files:
        # Skip Excel temporary lock files (~$Foo.xlsx)
        if xlsx_path.name.startswith("~$"):
            continue
        stem = xlsx_path.stem
        cfg = config_by_file.get(stem)
        try:
            records, skipped = process_file(xlsx_path, cfg)
            lab_name = cfg["lab_name"] if cfg else stem
            print(f"  ✓ {xlsx_path.name:25s} → {lab_name:20s} {len(records):5d} inregistrari "
                  f"({skipped} sarite)")
            all_records.extend(records)
            summary.append((lab_name, len(records)))
        except Exception as e:
            print(f"  ✗ {xlsx_path.name:25s} → EROARE: {e}", file=sys.stderr)
            sys.exit(1)

    # Scrie JSON-ul
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT, "w", encoding="utf-8") as f:
        json.dump(all_records, f, ensure_ascii=False, separators=(",", ":"))

    print(f"\n=== Rezumat ===")
    for lab, n in summary:
        print(f"  {lab:25s} {n:5d}")
    print(f"  {'TOTAL':25s} {len(all_records):5d}")
    print(f"\n→ Scris in: {OUTPUT.relative_to(ROOT)} ({OUTPUT.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
